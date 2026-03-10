"""
Contractor Import Service — Parse Excel files and create Prospect records.

Handles:
- Parsing the Dallas registered contractors Excel format
- Name normalization (LAST, FIRST → First Last)
- City-State-Zip splitting
- Email deduplication
- Registration type → industry tag mapping
"""

import io
import logging
import re
import uuid
from datetime import datetime, timezone

from sqlalchemy import select, func

from api.database import async_session_factory
from api.models.prospect import Prospect

logger = logging.getLogger("outreach.import")

# ── Registration Type → Industry Tag + Noun Mapping ──────────────────
REGISTRATION_MAP = {
    "Backflow (BF)":                    ("backflow", "backflow technician"),
    "Building (BU)":                    ("general_contractor", "general contractor"),
    "Demolition (DE)":                  ("demolition", "demolition contractor"),
    "Electrical (EL)":                  ("electrician", "electrician"),
    "Electrical Sign (ES)":             ("sign_contractor", "sign contractor"),
    "Energy Code (EC)":                 ("energy_consultant", "energy consultant"),
    "Fence (FE)":                       ("fence_contractor", "fence contractor"),
    "Fire Alarm (FA)":                  ("fire_alarm", "fire alarm technician"),
    "Fire Sprinkler (Minor Work) (FS)": ("fire_sprinkler", "fire sprinkler contractor"),
    "Foundation (FO)":                  ("foundation", "foundation contractor"),
    "Green Building Phase 2 (2G)":      ("green_building", "green building contractor"),
    "Green Building Provider (GB)":     ("green_building", "green building provider"),
    "Landscape (LA)":                   ("landscaper", "landscaper"),
    "Lawn Sprinkler (LS)":              ("irrigation", "irrigation contractor"),
    "Mechanical (ME)":                  ("hvac", "HVAC contractor"),
    "Medical Gas (MG)":                 ("medical_gas", "medical gas technician"),
    "Moving (MO)":                      ("mover", "moving company"),
    "Paving (Sidewalk, Drive Approaches) (PV)": ("paving", "paving contractor"),
    "Plumbing (PL)":                    ("plumber", "plumber"),
    "Residential General (BR)":         ("residential_builder", "residential builder"),
    "Roofing (RO)":                     ("roofer", "roofer"),
    "Swimming Pool (SW)":               ("pool_contractor", "pool contractor"),
    "Tree Service (TS)":                ("tree_service", "tree service"),
    "Water Treatment (WT)":             ("water_treatment", "water treatment specialist"),
}


def _parse_name(raw: str) -> tuple[str, str]:
    """
    Parse 'LAST, FIRST MIDDLE' → (first_name, full_name).
    Returns (first_name, "First Last") or ("", raw) if parsing fails.
    """
    raw = raw.strip()
    if "," in raw:
        parts = raw.split(",", 1)
        last = parts[0].strip().title()
        first_parts = parts[1].strip().split()
        first = first_parts[0].title() if first_parts else ""
        return first, f"{first} {last}".strip()
    # No comma — just title-case the whole thing
    return raw.split()[0].title() if raw.split() else "", raw.title()


def _parse_city_state(raw: str) -> tuple[str, str, str]:
    """
    Parse 'DALLAS ,TX 75243' → (city, state, zip).
    Handles various spacing/formatting inconsistencies.
    """
    raw = raw.strip()
    # Try pattern: CITY ,ST ZIP or CITY, ST ZIP
    m = re.match(r'^(.+?)\s*,\s*([A-Z]{2})\s+(\d{5}(?:-\d{4})?)$', raw, re.IGNORECASE)
    if m:
        return m.group(1).strip().title(), m.group(2).upper(), m.group(3)
    # Try: CITY ,ST (no zip)
    m = re.match(r'^(.+?)\s*,\s*([A-Z]{2})\s*$', raw, re.IGNORECASE)
    if m:
        return m.group(1).strip().title(), m.group(2).upper(), ""
    return raw.title(), "TX", ""


def _normalize_phone(raw: str) -> str:
    """Normalize phone to (XXX) XXX-XXXX format."""
    digits = re.sub(r'\D', '', raw)
    if len(digits) == 10:
        return f"({digits[:3]}) {digits[3:6]}-{digits[6:]}"
    if len(digits) == 11 and digits[0] == '1':
        return f"({digits[1:4]}) {digits[4:7]}-{digits[7:]}"
    return raw.strip()


def parse_contractors_excel(file_bytes: bytes) -> dict:
    """
    Parse the Dallas registered contractors Excel file.
    Returns {
        "rows": [parsed dicts],
        "total": int,
        "with_email": int,
        "skipped_no_email": int,
        "duplicates": int,
        "registration_types": [str],
        "errors": [str],
    }
    """
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True)
    ws = wb.active or wb[wb.sheetnames[0]]

    rows = []
    seen_emails = {}
    errors = []
    skipped_no_email = 0
    duplicates = 0
    reg_types = set()
    row_num = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        row_num += 1
        if not row or not row[1]:
            continue

        reg_type = str(row[0] or "").strip()
        contractor_raw = str(row[1] or "").strip()
        address = str(row[2] or "").strip()
        city_state_raw = str(row[3] or "").strip()
        phone_raw = str(row[4] or "").strip()
        email_raw = str(row[5] or "").strip().lower()

        if not email_raw:
            skipped_no_email += 1
            continue

        reg_types.add(reg_type)

        first_name, full_name = _parse_name(contractor_raw)
        city, state, zip_code = _parse_city_state(city_state_raw)
        phone = _normalize_phone(phone_raw) if phone_raw else ""

        # Industry mapping
        tag_info = REGISTRATION_MAP.get(reg_type)
        industry_tag = tag_info[0] if tag_info else "contractor"
        industry_noun = tag_info[1] if tag_info else "contractor"

        # Deduplicate by email
        if email_raw in seen_emails:
            # Merge registration types
            existing = seen_emails[email_raw]
            existing_tags = existing.get("tags", [])
            if reg_type not in existing_tags:
                existing_tags.append(reg_type)
                existing["tags"] = existing_tags
            duplicates += 1
            continue

        parsed = {
            "owner_name": full_name,
            "owner_first": first_name,
            "owner_email": email_raw,
            "business_name": full_name,  # Contractor name as business name
            "business_type": industry_tag,
            "industry_noun": industry_noun,
            "address": address.title() if address else "",
            "city": city,
            "state": state,
            "zip": zip_code,
            "phone": phone,
            "tags": [reg_type],
            "registration_type": reg_type,
        }
        seen_emails[email_raw] = parsed
        rows.append(parsed)

    wb.close()

    return {
        "rows": rows,
        "total": row_num,
        "with_email": len(rows),
        "skipped_no_email": skipped_no_email,
        "duplicates": duplicates,
        "registration_types": sorted(reg_types),
        "errors": errors,
    }


async def import_prospects(rows: list[dict], source: str = "contractor_registry") -> dict:
    """
    Insert parsed rows as Prospect records. Skips existing emails.
    Returns {"created": int, "skipped_existing": int, "errors": int}.
    """
    created = 0
    skipped = 0
    errors = 0

    async with async_session_factory() as db:
        # Get all existing emails for fast dedup
        result = await db.execute(
            select(func.lower(Prospect.owner_email))
            .where(Prospect.owner_email.isnot(None))
        )
        existing_emails = {e for (e,) in result.all()}

        for row in rows:
            email = row["owner_email"].lower()
            if email in existing_emails:
                skipped += 1
                continue

            try:
                prospect = Prospect(
                    id=uuid.uuid4(),
                    business_name=row["business_name"],
                    business_type=row["business_type"],
                    industry_tag=row.get("business_type"),
                    address=row.get("address"),
                    city=row["city"],
                    state=row["state"],
                    zip=row.get("zip"),
                    phone=row.get("phone"),
                    owner_name=row["owner_name"],
                    owner_email=email,
                    email_source="contractor_registry_import",
                    status="imported",
                    source=source,
                    tags=row.get("tags", []),
                    has_website=None,  # unknown
                    created_at=datetime.now(timezone.utc),
                )
                db.add(prospect)
                existing_emails.add(email)
                created += 1

                if created % 500 == 0:
                    await db.flush()
                    logger.info("Import progress: %d created so far", created)

            except Exception as e:
                errors += 1
                logger.warning("Failed to create prospect for %s: %s", email, e)

        await db.commit()

    logger.info("Import complete: %d created, %d skipped (existing), %d errors",
                created, skipped, errors)
    return {"created": created, "skipped_existing": skipped, "errors": errors}
